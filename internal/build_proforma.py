"""Build 1601-central-proforma.xlsx from assumptions.json.

Phase Zero pro forma lite — directional, range-based, confidence-labeled.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

HERE = Path(__file__).parent
ASSUMP_PATH = HERE / "assumptions.json"
OUT_PATH = HERE.parent / "deliverables" / "1601-central-proforma.xlsx"

with open(ASSUMP_PATH) as f:
    A = json.load(f)

# ============ STYLES ============
FONT_BODY = Font(name="Arial", size=10)
FONT_BODY_BOLD = Font(name="Arial", size=10, bold=True)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="14322A")
FONT_SUBTITLE = Font(name="Arial", size=11, italic=True, color="6B6458")
FONT_INPUT = Font(name="Arial", size=10, color="0000FF")        # blue = hardcoded input
FONT_FORMULA = Font(name="Arial", size=10, color="000000")      # black = formula
FONT_CROSSREF = Font(name="Arial", size=10, color="008000")     # green = cross-sheet
FONT_NOTE = Font(name="Arial", size=9, italic=True, color="6B6458")

FILL_HEADER = PatternFill("solid", start_color="14322A")        # forest
FILL_SUBHEADER = PatternFill("solid", start_color="EBE4D3")     # paper deep
FILL_UNCONFIRMED = PatternFill("solid", start_color="FFFF00")   # yellow per xlsx std
FILL_KEYROW = PatternFill("solid", start_color="F4EFE4")        # paper
FILL_TOTAL = PatternFill("solid", start_color="D4CDB8")         # rule soft

THIN = Side(border_style="thin", color="BEB7A5")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM = Border(bottom=Side(border_style="medium", color="13110E"))

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

FMT_USD = '"$"#,##0;("$"#,##0);"-"'
FMT_USD_CENTS = '"$"#,##0.00;("$"#,##0.00);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_NUM = '#,##0;(#,##0);"-"'
FMT_NUM1 = '#,##0.0;(#,##0.0);"-"'
FMT_MULTIPLE = '0.00"x"'

# ============ BUILD WORKBOOK ============
wb = Workbook()

# Delete default sheet; we'll build ours
del wb["Sheet"]


def style_title(ws, row, title, subtitle=None):
    ws.cell(row=row, column=1, value=title).font = FONT_TITLE
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = FONT_SUBTITLE
    return row + (3 if subtitle else 2)


def write_header(ws, row, headers, widths=None):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    if widths:
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w


def add_defined_name(wb, name, sheet, cell_ref):
    """Workbook-scoped named range pointing at an absolute cell."""
    ref = f"'{sheet}'!${cell_ref.split('!')[-1].replace('$','')}"
    # Build proper absolute ref
    col, rownum = "", ""
    for ch in cell_ref:
        if ch.isalpha():
            col += ch
        elif ch.isdigit():
            rownum += ch
    ref = f"'{sheet}'!${col}${rownum}"
    defn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = defn


# ============================================================
# TAB 1: README
# ============================================================
readme = wb.create_sheet("README")
readme.sheet_view.showGridLines = False

r = style_title(readme, 1, "1601–1607 Central Ave · Phase Zero Pro Forma Lite",
                "Directional, range-based, confidence-labeled. Not lender-grade.")
r += 1

readme.cell(row=r, column=1, value="Project").font = FONT_BODY_BOLD
readme.cell(row=r, column=2, value=A["meta"]["project"]).font = FONT_BODY
r += 1
readme.cell(row=r, column=1, value="Version").font = FONT_BODY_BOLD
readme.cell(row=r, column=2, value=A["meta"]["version"]).font = FONT_BODY
r += 1
readme.cell(row=r, column=1, value="As of").font = FONT_BODY_BOLD
readme.cell(row=r, column=2, value=A["meta"]["as_of"]).font = FONT_BODY
r += 1
readme.cell(row=r, column=1, value="Source of truth").font = FONT_BODY_BOLD
readme.cell(row=r, column=2, value="internal/assumptions.json").font = FONT_BODY
r += 2

readme.cell(row=r, column=1, value="Status Legend").font = FONT_TITLE
r += 1
status_legend = [
    ("confirmed",   "Parcel and ownership facts. Auditor / CAGIS sourced."),
    ("derived",     "Calculated from confirmed inputs. Not independently verified."),
    ("indicative",  "Industry-standard pricing / economic assumption. Not a quote."),
    ("target",      "A design target for the project scope (not a forecast)."),
    ("unconfirmed", "Requires outside validation: zoning, incentive, tract eligibility, or similar."),
]
for label, desc in status_legend:
    c1 = readme.cell(row=r, column=1, value=label)
    c1.font = FONT_BODY_BOLD
    if label == "unconfirmed":
        c1.fill = FILL_UNCONFIRMED
    readme.cell(row=r, column=2, value=desc).font = FONT_BODY
    r += 1
r += 1

readme.cell(row=r, column=1, value="Principles").font = FONT_TITLE
r += 1
for p in A["meta"]["principles"]:
    readme.cell(row=r, column=1, value="•").font = FONT_BODY
    readme.cell(row=r, column=2, value=p).font = FONT_BODY
    r += 1
r += 1

readme.cell(row=r, column=1, value="How to use").font = FONT_TITLE
r += 1
usage = [
    "All tweakable inputs live on the Inputs tab. Change a value there and every tab recalcs.",
    "Named ranges follow PathA_*, PathB_*, Stack_*, etc. — visible via Formulas → Name Manager.",
    "Blue text = hardcoded input (change these). Black = formula. Green = cross-sheet link.",
    "Yellow-highlighted rows are unconfirmed assumptions requiring diligence.",
    "The ‘Value Gap’ is TDC minus stabilized value. The ‘Capital Gap’ is TDC minus sum of sources. Keep them separate.",
    "Waterfall v1 shows directional economics only. IRR is not modeled — add hold period / cash flow timing first.",
]
for u in usage:
    readme.cell(row=r, column=1, value="•").font = FONT_BODY
    readme.cell(row=r, column=2, value=u).font = FONT_BODY
    r += 1

readme.column_dimensions["A"].width = 16
readme.column_dimensions["B"].width = 95

# ============================================================
# TAB 2: Inputs — single source of truth
# ============================================================
inp = wb.create_sheet("Inputs")
inp.sheet_view.showGridLines = False

r = style_title(inp, 1, "Inputs · Single Source of Truth",
                "One row per assumption. Named ranges point at the Base column.")
r += 1

HEADERS = ["Category", "Key (Name)", "Label", "Unit", "Status", "Low", "Base", "High", "Source", "Notes"]
WIDTHS = [14, 28, 38, 14, 14, 14, 14, 14, 36, 40]
write_header(inp, r, HEADERS, WIDTHS)
HDR_ROW = r
r += 1

# Rows: (category, key, label, unit, status, low, base, high, source, notes)
ROWS = [
    # site
    ("site", "marvin_land_sf",         "Smith Marvin land — total",               "sf",  "confirmed",   None,   6010,    None,   "HCAuditor · 3 parcels", ""),
    ("site", "marvin_land_ac",         "Smith Marvin land — acres",               "ac",  "confirmed",   None,   0.138,   None,   "HCAuditor · 3 parcels", ""),
    ("site", "marvin_assessed_value",  "Smith Marvin land — assessed value",      "usd", "confirmed",   None,   31640,   None,   "HCAuditor", ""),
    ("site", "public_adjacent_ac",     "Public adjacent (City + Land Bank)",      "ac",  "confirmed",   None,   0.302,   None,   "CAGIS", ""),
    ("site", "distressed_adjacent_ac", "Distressed adjacent (Medina + SW Ohio)",  "ac",  "confirmed",   None,   0.258,   None,   "CAGIS", ""),
    ("site", "assemblage_target_sf",   "Path B assemblage — target",              "sf",  "derived",     None,   30400,   None,   "Sum of targeted parcels", ""),
    # zoning / incentives (non-numeric)
    ("zoning",     "zoning_base",             "Zoning base",                  "text", "confirmed",   None, "MG-T", None, "CAGIS", "Manufacturing General — Transect overlay"),
    ("zoning",     "hotel_entitlement_path",  "Hotel entitlement path",       "text", "unconfirmed", None, "Planning confirmation required", None, "Cincinnati DSD", "CUP vs map amendment TBD"),
    ("incentives", "commercial_cra_status",   "Commercial CRA",               "text", "unconfirmed", None, "Negotiated / unconfirmed",       None, "ChooseCincy — commercial CRAs", "Lift tier is residential; commercial is negotiated"),
    ("incentives", "nmtc_status",             "NMTC eligibility",             "text", "unconfirmed", None, "Tract + allocation unconfirmed", None, "CDFI Fund tract lookup", ""),
    ("incentives", "tif_status",              "TIF status",                   "text", "unconfirmed", None, "Parcel-level check required",    None, "Cincinnati TIF district maps", ""),
    # path A
    ("path_a", "keys",             "Path A · keys",              "rooms",       "target",     24,      26,      28,      "Capacity from micro-hotel footprint", ""),
    ("path_a", "gsf",              "Path A · gross sf",          "gsf",         "indicative", None,    24000,   None,    "6 floors × ~4k plate", ""),
    ("path_a", "hard_cost_psf",    "Path A · hard cost per sf",  "usd_per_gsf", "indicative", 275,     300,     325,     "RSMeans Cincinnati 2026 Type III", ""),
    ("path_a", "soft_cost_pct",    "Path A · soft cost %",       "pct",         "indicative", 0.15,    0.17,    0.20,    "Industry", "A/E, legal, financing, permits"),
    ("path_a", "ffe_per_key",      "Path A · FF&E per key",      "usd_per_key", "indicative", 15000,   20000,   25000,   "Industry", "Micro-hotel spec"),
    ("path_a", "land_value",       "Path A · land value",        "usd",         "indicative", 500000,  700000,  800000,  "6010 sf at corner premium", "Appraisal required"),
    ("path_a", "contingency_pct",  "Path A · contingency %",     "pct",         "indicative", None,    0.10,    None,    "Industry standard", ""),
    ("path_a", "occupancy",        "Path A · occupancy (stab.)", "pct",         "indicative", 0.65,    0.70,    0.75,    "Micro-hotel stabilized y3", ""),
    ("path_a", "adr",              "Path A · ADR",               "usd",         "indicative", 120,     130,     140,     "West End comp read", "STR market study required"),
    ("path_a", "stabilized_noi",   "Path A · stabilized NOI",    "usd",         "indicative", 350000,  425000,  500000,  "Top-down sanity check", "Not derived from ADR×occ in v1"),
    ("path_a", "cap_rate",         "Path A · exit cap rate",     "pct",         "indicative", 0.080,   0.085,   0.090,   "Cincinnati select-service boutique", ""),
    # path B
    ("path_b", "keys",             "Path B · keys",              "rooms",       "target",     60,      70,      80,      "Assemblage scale", ""),
    ("path_b", "gsf",              "Path B · gross sf",          "gsf",         "indicative", None,    65000,   None,    "8 floors × ~8k plate", ""),
    ("path_b", "hard_cost_psf",    "Path B · hard cost per sf",  "usd_per_gsf", "indicative", None,    300,     None,    "RSMeans Cincinnati 2026", "Could flex to Type I if >6 stories"),
    ("path_b", "soft_cost_pct",    "Path B · soft cost %",       "pct",         "indicative", None,    0.18,    None,    "Industry", ""),
    ("path_b", "ffe_per_key",      "Path B · FF&E per key",      "usd_per_key", "indicative", None,    20000,   None,    "Industry", "Boutique tier"),
    ("path_b", "land_acquisition", "Path B · land acquisition",  "usd",         "indicative", 1500000, 2000000, 2500000, "All parcels incl. Marvin", "Public land nominal; Medina distressed"),
    ("path_b", "contingency_pct",  "Path B · contingency %",     "pct",         "indicative", None,    0.10,    None,    "Industry", ""),
    ("path_b", "occupancy",        "Path B · occupancy (stab.)", "pct",         "indicative", 0.65,    0.70,    0.75,    "Boutique stabilized y3", ""),
    ("path_b", "adr",              "Path B · ADR",               "usd",         "indicative", 140,     160,     180,     "Soft-brand boutique", "STR study required"),
    ("path_b", "stabilized_noi",   "Path B · stabilized NOI",    "usd",         "indicative", 1800000, 2000000, 2200000, "Top-down sanity check", "Not derived from ADR×occ in v1"),
    ("path_b", "cap_rate",         "Path B · exit cap rate",     "pct",         "indicative", None,    0.080,   None,    "Cincinnati boutique", ""),
    # capital stack
    ("stack", "senior_debt",              "Senior construction debt",   "usd", "indicative",  16000000, 16500000, 17000000, "55% LTC · CRA-motivated lender", ""),
    ("stack", "nmtc_equity",              "NMTC equity",                "usd", "unconfirmed", 3000000,  3500000,  4000000,  "Via CDE allocation",            "Only if tract + allocation confirmed"),
    ("stack", "htc_equity",               "Historic Tax Credit equity", "usd", "unconfirmed", 0,        0,        1000000,  "Ohio HTC",                      "Only if preserving structure"),
    ("stack", "city_gap",                 "City of Cincinnati gap",     "usd", "unconfirmed", 1000000,  1500000,  2000000,  "DSD commercial gap fund",       "Negotiated"),
    ("stack", "public_land_contribution", "Public land (implied)",      "usd", "unconfirmed", 500000,   750000,   1000000,  "City + Land Bank at nominal",   "Negotiated disposition"),
    ("stack", "marvin_land_contribution", "Marvin land (implied)",      "usd", "indicative",  500000,   700000,   800000,   "Appraised FMV",                 "Sale or contribution"),
    ("stack", "sponsor_equity",           "Sponsor (GP) equity",        "usd", "indicative",  1500000,  1750000,  2000000,  "Developer skin in the game",    ""),
    ("stack", "lp_equity",                "LP equity (passive)",        "usd", "indicative",  2000000,  2500000,  3000000,  "Local HNW check-writers",       ""),
    # pre-dev
    ("predev", "total_predev",      "Total pre-development spend", "usd",    "indicative", 600000, 850000, 1100000, "Line items across 15–18 mo", ""),
    ("predev", "timeline_months",   "Pre-development timeline",    "months", "indicative", 15,     16.5,   18,      "From option to close",        ""),
    # equity options
    ("options", "option2_equity_pct",     "Opt II — equity share to Marvin", "pct", "modeled", 0.10, 0.11, 0.12, "Land contribution / total equity", "Pari-passu"),
    ("options", "option3_pref_return_pct","Opt III — preferred return rate", "pct", "modeled", None, 0.08, None, "Common market pref",              ""),
    ("options", "option3_residual_pct",   "Opt III — residual equity share", "pct", "modeled", 0.03, 0.04, 0.05, "Back-end participation",           ""),
]

# Write rows, capture row numbers for each key so we can set defined names
KEY_ROW = {}  # (category, key) -> excel row number

for row_data in ROWS:
    cat, key, label, unit, status, low, base, high, source, notes = row_data
    inp.cell(row=r, column=1, value=cat).font = FONT_BODY
    inp.cell(row=r, column=2, value=f"{cat}.{key}").font = FONT_BODY_BOLD
    inp.cell(row=r, column=3, value=label).font = FONT_BODY
    inp.cell(row=r, column=4, value=unit).font = FONT_BODY

    status_c = inp.cell(row=r, column=5, value=status)
    status_c.font = FONT_BODY_BOLD if status == "unconfirmed" else FONT_BODY
    if status == "unconfirmed":
        status_c.fill = FILL_UNCONFIRMED

    # numeric columns
    low_c = inp.cell(row=r, column=6, value=low)
    base_c = inp.cell(row=r, column=7, value=base)
    high_c = inp.cell(row=r, column=8, value=high)

    # format: pct / usd / number / text
    if unit == "pct":
        for c in (low_c, base_c, high_c):
            c.number_format = FMT_PCT
            c.font = FONT_INPUT if c.value is not None else FONT_BODY
    elif unit == "usd" or unit.startswith("usd_"):
        for c in (low_c, base_c, high_c):
            c.number_format = FMT_USD
            c.font = FONT_INPUT if c.value is not None else FONT_BODY
    elif unit == "text":
        # base column holds text value
        base_c.font = FONT_INPUT if isinstance(base_c.value, str) else FONT_BODY
    else:  # sf, ac, rooms, gsf, months
        for c in (low_c, base_c, high_c):
            c.number_format = FMT_NUM1 if unit == "months" else FMT_NUM
            c.font = FONT_INPUT if c.value is not None else FONT_BODY

    for c in (low_c, base_c, high_c):
        c.alignment = ALIGN_RIGHT

    inp.cell(row=r, column=9, value=source).font = FONT_NOTE
    inp.cell(row=r, column=10, value=notes).font = FONT_NOTE

    KEY_ROW[(cat, key)] = r
    r += 1

# Freeze header
inp.freeze_panes = "A" + str(HDR_ROW + 1)

# ============================================================
# Define named ranges — each points at the "Base" cell (column G, col 7)
# ============================================================
# Naming convention: <CategoryCamel>_<Key> (e.g., PathA_Keys, Stack_SeniorDebt)
CAT_PREFIX = {
    "site": "Site",
    "zoning": "Zoning",
    "incentives": "Incent",
    "path_a": "PathA",
    "path_b": "PathB",
    "stack": "Stack",
    "predev": "PreDev",
    "options": "Opt",
}

def key_to_camel(k):
    parts = k.split("_")
    return "".join(p.capitalize() for p in parts)


NAMES = {}  # semantic_name -> (sheet, cell)
for (cat, key), row_num in KEY_ROW.items():
    prefix = CAT_PREFIX[cat]
    name = f"{prefix}_{key_to_camel(key)}"
    cell_ref = f"G{row_num}"
    NAMES[name] = ("Inputs", cell_ref)
    defn = DefinedName(name=name, attr_text=f"'Inputs'!$G${row_num}")
    wb.defined_names[name] = defn


# Convenience handle for writing formulas
def n(name):
    """Returns the named-range literal for use in formulas."""
    return name


# ============================================================
# TAB 3: Path A
# ============================================================
pa = wb.create_sheet("Path A - Micro")
pa.sheet_view.showGridLines = False
r = style_title(pa, 1, "Path A · Micro-Hotel on Marvin's Land",
                "24–28 keys · Type III wood over podium · Ollie's on ground floor")
r += 1

pa.cell(row=r, column=1, value="Line").font = FONT_HEADER
pa.cell(row=r, column=1).fill = FILL_HEADER
pa.cell(row=r, column=2, value="Formula").font = FONT_HEADER
pa.cell(row=r, column=2).fill = FILL_HEADER
pa.cell(row=r, column=3, value="Value").font = FONT_HEADER
pa.cell(row=r, column=3).fill = FILL_HEADER
for col in (1, 2, 3):
    pa.cell(row=r, column=col).alignment = ALIGN_CENTER
    pa.cell(row=r, column=col).border = BORDER_ALL
r += 1

def row_line(ws, r, label, formula, desc, fmt=FMT_USD, is_total=False):
    c1 = ws.cell(row=r, column=1, value=label)
    c2 = ws.cell(row=r, column=2, value=desc)
    c3 = ws.cell(row=r, column=3, value=formula)
    c1.font = FONT_BODY_BOLD if is_total else FONT_BODY
    c2.font = FONT_NOTE
    c3.font = FONT_CROSSREF
    c3.number_format = fmt
    c3.alignment = ALIGN_RIGHT
    if is_total:
        for c in (c1, c2, c3):
            c.fill = FILL_TOTAL
    return r + 1

r = row_line(pa, r, "Keys",              f"={n('PathA_Keys')}",           "PathA_Keys", fmt=FMT_NUM)
r = row_line(pa, r, "GSF",               f"={n('PathA_Gsf')}",            "PathA_Gsf", fmt=FMT_NUM)
r = row_line(pa, r, "Hard cost / sf",    f"={n('PathA_HardCostPsf')}",    "PathA_HardCostPsf", fmt=FMT_USD)
r += 1
r = row_line(pa, r, "Hard cost",         f"={n('PathA_Gsf')}*{n('PathA_HardCostPsf')}", "GSF × $/sf")
hard_cost_row = r - 1
r = row_line(pa, r, "Soft cost",         f"=C{hard_cost_row}*{n('PathA_SoftCostPct')}", "Hard × soft %")
soft_cost_row = r - 1
r = row_line(pa, r, "FF&E",              f"={n('PathA_Keys')}*{n('PathA_FfePerKey')}", "Keys × $/key")
ffe_row = r - 1
r = row_line(pa, r, "Land value",        f"={n('PathA_LandValue')}",      "Marvin parcels FMV")
land_row = r - 1
r = row_line(pa, r, "Contingency",       f"=(C{hard_cost_row}+C{soft_cost_row}+C{ffe_row}+C{land_row})*{n('PathA_ContingencyPct')}", "10% of subtotal")
cont_row = r - 1
r = row_line(pa, r, "Total Development Cost (TDC)",
             f"=C{hard_cost_row}+C{soft_cost_row}+C{ffe_row}+C{land_row}+C{cont_row}", "Sum", is_total=True)
tdc_row = r - 1
r += 1

r = row_line(pa, r, "Stabilized NOI",    f"={n('PathA_StabilizedNoi')}",  "Top-down assumption")
noi_row = r - 1
r = row_line(pa, r, "Yield on Cost",     f"=C{noi_row}/C{tdc_row}",       "NOI / TDC", fmt=FMT_PCT)
r = row_line(pa, r, "Exit Cap Rate",     f"={n('PathA_CapRate')}",        "Market read", fmt=FMT_PCT)
cap_row = r - 1
r = row_line(pa, r, "Stabilized Value",  f"=C{noi_row}/C{cap_row}",       "NOI / Cap", is_total=True)
val_row = r - 1
r = row_line(pa, r, "Value Gap",         f"=C{tdc_row}-C{val_row}",       "TDC − Value (positive = underwater)", is_total=True)
r += 1

# Revenue sanity check
pa.cell(row=r, column=1, value="Revenue sanity check (informational)").font = FONT_SUBTITLE
r += 1
r = row_line(pa, r, "ADR",               f"={n('PathA_Adr')}",            "Avg Daily Rate", fmt=FMT_USD)
r = row_line(pa, r, "Occupancy",         f"={n('PathA_Occupancy')}",      "Stabilized", fmt=FMT_PCT)
r = row_line(pa, r, "Implied room rev",  f"={n('PathA_Keys')}*365*{n('PathA_Occupancy')}*{n('PathA_Adr')}", "Keys × 365 × Occ × ADR")

pa.column_dimensions["A"].width = 38
pa.column_dimensions["B"].width = 42
pa.column_dimensions["C"].width = 18

# ============================================================
# TAB 4: Path B
# ============================================================
pb = wb.create_sheet("Path B - Assemblage")
pb.sheet_view.showGridLines = False
r = style_title(pb, 1, "Path B · Assemblage Play",
                "60–80 keys · matches the flyer's rendering · requires adjacent parcels")
r += 1

pb.cell(row=r, column=1, value="Line").font = FONT_HEADER
pb.cell(row=r, column=1).fill = FILL_HEADER
pb.cell(row=r, column=2, value="Formula").font = FONT_HEADER
pb.cell(row=r, column=2).fill = FILL_HEADER
pb.cell(row=r, column=3, value="Value").font = FONT_HEADER
pb.cell(row=r, column=3).fill = FILL_HEADER
for col in (1, 2, 3):
    pb.cell(row=r, column=col).alignment = ALIGN_CENTER
    pb.cell(row=r, column=col).border = BORDER_ALL
r += 1

r = row_line(pb, r, "Keys",              f"={n('PathB_Keys')}",           "PathB_Keys", fmt=FMT_NUM)
r = row_line(pb, r, "GSF",               f"={n('PathB_Gsf')}",            "PathB_Gsf", fmt=FMT_NUM)
r = row_line(pb, r, "Hard cost / sf",    f"={n('PathB_HardCostPsf')}",    "PathB_HardCostPsf", fmt=FMT_USD)
r += 1
r = row_line(pb, r, "Hard cost",         f"={n('PathB_Gsf')}*{n('PathB_HardCostPsf')}", "GSF × $/sf")
hc = r - 1
r = row_line(pb, r, "Soft cost",         f"=C{hc}*{n('PathB_SoftCostPct')}", "Hard × soft %")
sc = r - 1
r = row_line(pb, r, "FF&E",              f"={n('PathB_Keys')}*{n('PathB_FfePerKey')}", "Keys × $/key")
ffe = r - 1
r = row_line(pb, r, "Land acquisition",  f"={n('PathB_LandAcquisition')}","All parcels incl. Marvin")
land = r - 1
r = row_line(pb, r, "Contingency",       f"=(C{hc}+C{sc}+C{ffe}+C{land})*{n('PathB_ContingencyPct')}", "10% of subtotal")
cont = r - 1
r = row_line(pb, r, "Total Development Cost (TDC)",
             f"=C{hc}+C{sc}+C{ffe}+C{land}+C{cont}", "Sum", is_total=True)
pb_tdc_row = r - 1
r += 1

r = row_line(pb, r, "Stabilized NOI",    f"={n('PathB_StabilizedNoi')}",  "Top-down assumption")
noi = r - 1
r = row_line(pb, r, "Yield on Cost",     f"=C{noi}/C{pb_tdc_row}",        "NOI / TDC", fmt=FMT_PCT)
r = row_line(pb, r, "Exit Cap Rate",     f"={n('PathB_CapRate')}",        "Market read", fmt=FMT_PCT)
cap = r - 1
r = row_line(pb, r, "Stabilized Value",  f"=C{noi}/C{cap}",               "NOI / Cap", is_total=True)
pb_val_row = r - 1
r = row_line(pb, r, "Value Gap",         f"=C{pb_tdc_row}-C{pb_val_row}", "TDC − Value", is_total=True)
r += 1

pb.cell(row=r, column=1, value="Revenue sanity check (informational)").font = FONT_SUBTITLE
r += 1
r = row_line(pb, r, "ADR",               f"={n('PathB_Adr')}",            "Avg Daily Rate", fmt=FMT_USD)
r = row_line(pb, r, "Occupancy",         f"={n('PathB_Occupancy')}",      "Stabilized", fmt=FMT_PCT)
r = row_line(pb, r, "Implied room rev",  f"={n('PathB_Keys')}*365*{n('PathB_Occupancy')}*{n('PathB_Adr')}", "Keys × 365 × Occ × ADR")

pb.column_dimensions["A"].width = 38
pb.column_dimensions["B"].width = 42
pb.column_dimensions["C"].width = 18

# Add a workbook-level named ranges for downstream sheets
wb.defined_names["PathB_TDC"] = DefinedName(name="PathB_TDC", attr_text=f"'Path B - Assemblage'!$C${pb_tdc_row}")
wb.defined_names["PathB_StabValue"] = DefinedName(name="PathB_StabValue", attr_text=f"'Path B - Assemblage'!$C${pb_val_row}")

# ============================================================
# TAB 5: Capital Stack
# ============================================================
cs = wb.create_sheet("Capital Stack")
cs.sheet_view.showGridLines = False
r = style_title(cs, 1, "Capital Stack · Path B Sources & Uses",
                "Sources (base case) vs. Path B TDC · gap = unfunded")
r += 1

cs.cell(row=r, column=1, value="Source").font = FONT_HEADER
cs.cell(row=r, column=1).fill = FILL_HEADER
cs.cell(row=r, column=2, value="Amount").font = FONT_HEADER
cs.cell(row=r, column=2).fill = FILL_HEADER
cs.cell(row=r, column=3, value="Status").font = FONT_HEADER
cs.cell(row=r, column=3).fill = FILL_HEADER
cs.cell(row=r, column=4, value="Notes").font = FONT_HEADER
cs.cell(row=r, column=4).fill = FILL_HEADER
for col in (1, 2, 3, 4):
    cs.cell(row=r, column=col).alignment = ALIGN_CENTER
    cs.cell(row=r, column=col).border = BORDER_ALL
r += 1

stack_items = [
    ("Senior construction debt",   "Stack_SeniorDebt",            "indicative",  "55% LTC · CRA-motivated bank"),
    ("NMTC equity",                "Stack_NmtcEquity",            "unconfirmed", "Tract + allocation unconfirmed"),
    ("Historic Tax Credit equity", "Stack_HtcEquity",             "unconfirmed", "Only if preserving structure"),
    ("City of Cincinnati gap",     "Stack_CityGap",               "unconfirmed", "DSD commercial gap fund"),
    ("Public land contribution",   "Stack_PublicLandContribution","unconfirmed", "City + Land Bank at nominal"),
    ("Marvin land contribution",   "Stack_MarvinLandContribution","indicative",  "Sale or equity contribution"),
    ("Sponsor (GP) equity",        "Stack_SponsorEquity",         "indicative",  "Developer skin in the game"),
    ("LP equity (passive)",        "Stack_LpEquity",              "indicative",  "Local HNW check-writers"),
]

first_src = r
for label, name, status, notes in stack_items:
    cs.cell(row=r, column=1, value=label).font = FONT_BODY
    c = cs.cell(row=r, column=2, value=f"={name}")
    c.font = FONT_CROSSREF
    c.number_format = FMT_USD
    c.alignment = ALIGN_RIGHT
    sc = cs.cell(row=r, column=3, value=status)
    sc.font = FONT_BODY_BOLD if status == "unconfirmed" else FONT_BODY
    if status == "unconfirmed":
        sc.fill = FILL_UNCONFIRMED
    cs.cell(row=r, column=4, value=notes).font = FONT_NOTE
    r += 1
last_src = r - 1

# Totals
cs.cell(row=r, column=1, value="Total Sources").font = FONT_BODY_BOLD
total_c = cs.cell(row=r, column=2, value=f"=SUM(B{first_src}:B{last_src})")
total_c.font = FONT_FORMULA
total_c.number_format = FMT_USD
total_c.alignment = ALIGN_RIGHT
for col in (1, 2, 3, 4):
    cs.cell(row=r, column=col).fill = FILL_TOTAL
total_sources_row = r
r += 1

cs.cell(row=r, column=1, value="Path B TDC (uses)").font = FONT_BODY_BOLD
tdc_ref = cs.cell(row=r, column=2, value="=PathB_TDC")
tdc_ref.font = FONT_CROSSREF
tdc_ref.number_format = FMT_USD
tdc_ref.alignment = ALIGN_RIGHT
cs.cell(row=r, column=3, value="from Path B tab").font = FONT_NOTE
pb_tdc_ref_row = r
r += 1

cs.cell(row=r, column=1, value="Capital Gap (TDC − Sources)").font = FONT_BODY_BOLD
gap = cs.cell(row=r, column=2, value=f"=B{pb_tdc_ref_row}-B{total_sources_row}")
gap.font = FONT_FORMULA
gap.number_format = FMT_USD
gap.alignment = ALIGN_RIGHT
cs.cell(row=r, column=4, value="Positive = unfunded · close with subsidy or structure").font = FONT_NOTE
for col in (1, 2, 3, 4):
    cs.cell(row=r, column=col).fill = FILL_TOTAL
capital_gap_row = r
r += 2

cs.cell(row=r, column=1, value="Note: Capital Gap and Value Gap are different. Value Gap = TDC − Stabilized Value (from Path B tab).").font = FONT_NOTE

cs.column_dimensions["A"].width = 32
cs.column_dimensions["B"].width = 18
cs.column_dimensions["C"].width = 14
cs.column_dimensions["D"].width = 50

# ============================================================
# TAB 6: Waterfall v1
# ============================================================
wf = wb.create_sheet("Waterfall v1")
wf.sheet_view.showGridLines = False
r = style_title(wf, 1, "Waterfall · Marvin's Three Equity Options (v1)",
                "Directional only. IRR not modeled — requires hold period, cash flow timing, sale costs.")
r += 1

wf.cell(row=r, column=1, value="Shared inputs").font = FONT_SUBTITLE
r += 1

# We'll model a hypothetical net sale proceeds as StabValue minus debt payoff minus sale costs.
wf.cell(row=r, column=1, value="Stabilized Value (Path B)").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=PathB_StabValue")
c.font = FONT_CROSSREF
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
stabv_row = r
r += 1

wf.cell(row=r, column=1, value="Less: senior debt payoff").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=Stack_SeniorDebt")
c.font = FONT_CROSSREF
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
debt_row = r
r += 1

wf.cell(row=r, column=1, value="Less: sale costs (3%)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=f"=B{stabv_row}*0.03")
c.font = FONT_FORMULA
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
salecost_row = r
r += 1

wf.cell(row=r, column=1, value="Net sale proceeds to equity").font = FONT_BODY_BOLD
c = wf.cell(row=r, column=2, value=f"=B{stabv_row}-B{debt_row}-B{salecost_row}")
c.font = FONT_FORMULA
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
for col in (1, 2):
    wf.cell(row=r, column=col).fill = FILL_TOTAL
net_proceeds_row = r
r += 2

# Option I
wf.cell(row=r, column=1, value="Option I · Land Sale + Anchor Tenant").font = FONT_TITLE
r += 1
wf.cell(row=r, column=1, value="Cash at closing (land sale)").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=Stack_MarvinLandContribution")
c.font = FONT_CROSSREF
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
r += 1
wf.cell(row=r, column=1, value="Back-end equity participation").font = FONT_BODY
c = wf.cell(row=r, column=2, value=0)
c.font = FONT_INPUT
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
r += 1
wf.cell(row=r, column=1, value="+ Below-market restaurant lease (separate P&L)").font = FONT_NOTE
r += 2

# Option II
wf.cell(row=r, column=1, value="Option II · Land as Equity (recommended)").font = FONT_TITLE
r += 1
wf.cell(row=r, column=1, value="Land contribution").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=Stack_MarvinLandContribution")
c.font = FONT_CROSSREF
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
r += 1
wf.cell(row=r, column=1, value="Equity share (pari-passu)").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=Opt_Option2EquityPct")
c.font = FONT_CROSSREF
c.number_format = FMT_PCT
c.alignment = ALIGN_RIGHT
opt2_pct_row = r
r += 1
wf.cell(row=r, column=1, value="Back-end value at sale").font = FONT_BODY_BOLD
c = wf.cell(row=r, column=2, value=f"=B{opt2_pct_row}*B{net_proceeds_row}")
c.font = FONT_FORMULA
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
for col in (1, 2):
    wf.cell(row=r, column=col).fill = FILL_TOTAL
r += 1
wf.cell(row=r, column=1, value="+ Below-market restaurant lease (separate P&L)").font = FONT_NOTE
r += 2

# Option III
wf.cell(row=r, column=1, value="Option III · Preferred Equity + Residual").font = FONT_TITLE
r += 1
wf.cell(row=r, column=1, value="Land contribution (pref basis)").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=Stack_MarvinLandContribution")
c.font = FONT_CROSSREF
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
pref_basis_row = r
r += 1
wf.cell(row=r, column=1, value="Pref return rate").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=Opt_Option3PrefReturnPct")
c.font = FONT_CROSSREF
c.number_format = FMT_PCT
c.alignment = ALIGN_RIGHT
pref_rate_row = r
r += 1
wf.cell(row=r, column=1, value="Annual preferred ($)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=f"=B{pref_basis_row}*B{pref_rate_row}")
c.font = FONT_FORMULA
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
r += 1
wf.cell(row=r, column=1, value="Residual equity share").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=Opt_Option3ResidualPct")
c.font = FONT_CROSSREF
c.number_format = FMT_PCT
c.alignment = ALIGN_RIGHT
residual_row = r
r += 1
wf.cell(row=r, column=1, value="Back-end residual at sale").font = FONT_BODY_BOLD
c = wf.cell(row=r, column=2, value=f"=B{residual_row}*B{net_proceeds_row}")
c.font = FONT_FORMULA
c.number_format = FMT_USD
c.alignment = ALIGN_RIGHT
for col in (1, 2):
    wf.cell(row=r, column=col).fill = FILL_TOTAL
r += 2

# ------------------ Upside Case (Path B) ------------------
r += 1
wf.cell(row=r, column=1, value="Upside Case · Path B at better assumptions + 7-yr hold").font = FONT_TITLE
r += 1
wf.cell(row=r, column=1, value="Shows where Option II materially outruns Option I. Uses NOI uplift, cap rate compression, and a holding period with NOI growth.").font = FONT_SUBTITLE
r += 2

# Inputs for upside scenario (hardcoded here, could be promoted to Inputs tab later)
wf.cell(row=r, column=1, value="NOI uplift vs. base (higher ADR + occ)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=0.15); c.font = FONT_INPUT; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT
uplift_row = r
r += 1

wf.cell(row=r, column=1, value="Cap rate compression (bps)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=-0.005); c.font = FONT_INPUT; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT
comp_row = r
r += 1

wf.cell(row=r, column=1, value="Hold period (years)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=7); c.font = FONT_INPUT; c.number_format = FMT_NUM; c.alignment = ALIGN_RIGHT
hold_row = r
r += 1

wf.cell(row=r, column=1, value="Annual NOI growth (held period)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=0.025); c.font = FONT_INPUT; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT
growth_row = r
r += 2

# Derived values
wf.cell(row=r, column=1, value="Upside stabilized NOI (Yr 1)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=f"=PathB_StabilizedNoi*(1+B{uplift_row})")
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
up_noi_row = r
r += 1

wf.cell(row=r, column=1, value="Upside exit cap rate").font = FONT_BODY
c = wf.cell(row=r, column=2, value=f"=PathB_CapRate+B{comp_row}")
c.font = FONT_FORMULA; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT
up_cap_row = r
r += 1

wf.cell(row=r, column=1, value="Year-N NOI (grown)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=f"=B{up_noi_row}*(1+B{growth_row})^B{hold_row}")
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
yrn_noi_row = r
r += 1

wf.cell(row=r, column=1, value="Year-N sale value").font = FONT_BODY_BOLD
c = wf.cell(row=r, column=2, value=f"=B{yrn_noi_row}/B{up_cap_row}")
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
yrn_val_row = r
for col in (1, 2):
    wf.cell(row=r, column=col).fill = FILL_TOTAL
r += 1

wf.cell(row=r, column=1, value="Less: senior debt payoff (assumes unchanged)").font = FONT_BODY
c = wf.cell(row=r, column=2, value="=Stack_SeniorDebt")
c.font = FONT_CROSSREF; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
up_debt_row = r
r += 1

wf.cell(row=r, column=1, value="Less: sale costs (3%)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=f"=B{yrn_val_row}*0.03")
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
up_sc_row = r
r += 1

wf.cell(row=r, column=1, value="Net sale proceeds to equity (upside)").font = FONT_BODY_BOLD
c = wf.cell(row=r, column=2, value=f"=B{yrn_val_row}-B{up_debt_row}-B{up_sc_row}")
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
for col in (1, 2):
    wf.cell(row=r, column=col).fill = FILL_TOTAL
up_net_row = r
r += 2

# Marvin's payout under each option — upside
wf.cell(row=r, column=1, value="Marvin — Option II (upside)").font = FONT_BODY_BOLD
c = wf.cell(row=r, column=2, value=f"=Opt_Option2EquityPct*B{up_net_row}")
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
for col in (1, 2):
    wf.cell(row=r, column=col).fill = FILL_TOTAL
r += 1

wf.cell(row=r, column=1, value="Marvin — Option III residual (upside)").font = FONT_BODY_BOLD
c = wf.cell(row=r, column=2, value=f"=Opt_Option3ResidualPct*B{up_net_row}")
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
for col in (1, 2):
    wf.cell(row=r, column=col).fill = FILL_TOTAL
r += 1

wf.cell(row=r, column=1, value="Option III cumulative pref (hold period)").font = FONT_BODY
c = wf.cell(row=r, column=2, value=f"=Stack_MarvinLandContribution*Opt_Option3PrefReturnPct*B{hold_row}")
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
r += 1

wf.cell(row=r, column=1, value="Option III total (pref + residual)").font = FONT_BODY_BOLD
opt3_total_formula = f"=Stack_MarvinLandContribution*Opt_Option3PrefReturnPct*B{hold_row}+Opt_Option3ResidualPct*B{up_net_row}"
c = wf.cell(row=r, column=2, value=opt3_total_formula)
c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
for col in (1, 2):
    wf.cell(row=r, column=col).fill = FILL_TOTAL
r += 2

# Side-by-side comparison
wf.cell(row=r, column=1, value="Comparison · Marvin's payout").font = FONT_TITLE
r += 1
wf.cell(row=r, column=1, value="Scenario").font = FONT_HEADER
wf.cell(row=r, column=1).fill = FILL_HEADER
wf.cell(row=r, column=2, value="Amount").font = FONT_HEADER
wf.cell(row=r, column=2).fill = FILL_HEADER
wf.cell(row=r, column=3, value="Timing").font = FONT_HEADER
wf.cell(row=r, column=3).fill = FILL_HEADER
for col in (1, 2, 3):
    wf.cell(row=r, column=col).alignment = ALIGN_CENTER
    wf.cell(row=r, column=col).border = BORDER_ALL
r += 1

compare_rows = [
    ("Option I (land sale, base case)",           "=Stack_MarvinLandContribution",                                      "At closing (Yr 0)"),
    ("Option II (11% equity, base case)",         f"=Opt_Option2EquityPct*B{net_proceeds_row}",                         "At sale (Yr ~5)"),
    ("Option II (11% equity, upside case)",       f"=Opt_Option2EquityPct*B{up_net_row}",                               f"At sale (Yr {7})"),
    ("Option III (pref + residual, upside)",      opt3_total_formula[1:],                                               f"Over {7} years + sale"),
]
for label, formula, timing in compare_rows:
    wf.cell(row=r, column=1, value=label).font = FONT_BODY
    c = wf.cell(row=r, column=2, value=f"={formula}" if not formula.startswith("=") else formula)
    c.font = FONT_FORMULA; c.number_format = FMT_USD; c.alignment = ALIGN_RIGHT
    wf.cell(row=r, column=3, value=timing).font = FONT_NOTE
    r += 1

r += 1

# Caveats
wf.cell(row=r, column=1, value="Caveats").font = FONT_TITLE
r += 1
caveats = [
    "Option I cash is at Yr 0; Option II/III are Yr 5–7. Not time-normalized — no discount rate or IRR applied.",
    "Upside case assumes the project appreciates past TDC. At base-case assumptions it does not.",
    "IRR not modeled. Requires annual cash flow timing, refinance proceeds, and distribution priority.",
    "Debt payoff held flat in upside. In practice, amortization + refi proceeds could change equity distributions.",
    "Sale costs held at 3% of gross; transfer tax, broker, legal may be higher.",
    "Commercial CRA abatement not yet modeled in NOI uplift — if granted, this is another upside lever.",
]
for c_text in caveats:
    wf.cell(row=r, column=1, value="•").font = FONT_NOTE
    wf.cell(row=r, column=2, value=c_text).font = FONT_NOTE
    r += 1

wf.column_dimensions["A"].width = 44
wf.column_dimensions["B"].width = 22
wf.column_dimensions["C"].width = 22

# ============================================================
# TAB 7: Pre-Dev Budget
# ============================================================
pd_ = wb.create_sheet("Pre-Dev Budget")
pd_.sheet_view.showGridLines = False
r = style_title(pd_, 1, "Pre-Development Budget",
                "15–18 months of at-risk spend before construction loan funds")
r += 1

pd_.cell(row=r, column=1, value="Activity").font = FONT_HEADER
pd_.cell(row=r, column=1).fill = FILL_HEADER
pd_.cell(row=r, column=2, value="Low").font = FONT_HEADER
pd_.cell(row=r, column=2).fill = FILL_HEADER
pd_.cell(row=r, column=3, value="Base").font = FONT_HEADER
pd_.cell(row=r, column=3).fill = FILL_HEADER
pd_.cell(row=r, column=4, value="High").font = FONT_HEADER
pd_.cell(row=r, column=4).fill = FILL_HEADER
pd_.cell(row=r, column=5, value="Timing (month)").font = FONT_HEADER
pd_.cell(row=r, column=5).fill = FILL_HEADER
for col in (1, 2, 3, 4, 5):
    pd_.cell(row=r, column=col).alignment = ALIGN_CENTER
    pd_.cell(row=r, column=col).border = BORDER_ALL
r += 1

predev_items = [
    ("Land options & earnest deposits",                  100000, 125000, 150000, "1–3"),
    ("Environmental Phase I + II",                        30000,  52000,  75000, "2–4"),
    ("Concept architecture + site planning",             100000, 150000, 200000, "3–8"),
    ("Zoning / CUP / rezoning legal & filings",           50000,  75000, 100000, "4–12"),
    ("Civil engineering / survey",                        40000,  57000,  75000, "4–9"),
    ("Hotel market study + operator interest memo",       40000,  50000,  60000, "2–5"),
    ("Capital stack structuring (NMTC, CRA, HTC legal)",  75000, 112000, 150000, "6–15"),
    ("Project management + carrying costs",              150000, 225000, 300000, "ongoing"),
]
first = r
for activity, lo, base_v, hi, timing in predev_items:
    pd_.cell(row=r, column=1, value=activity).font = FONT_BODY
    for col_idx, val in zip((2, 3, 4), (lo, base_v, hi)):
        c = pd_.cell(row=r, column=col_idx, value=val)
        c.font = FONT_INPUT
        c.number_format = FMT_USD
        c.alignment = ALIGN_RIGHT
    pd_.cell(row=r, column=5, value=timing).font = FONT_BODY
    pd_.cell(row=r, column=5).alignment = ALIGN_CENTER
    r += 1
last = r - 1

pd_.cell(row=r, column=1, value="Total Pre-Development").font = FONT_BODY_BOLD
for col_idx, letter in zip((2, 3, 4), ("B", "C", "D")):
    c = pd_.cell(row=r, column=col_idx, value=f"=SUM({letter}{first}:{letter}{last})")
    c.font = FONT_FORMULA
    c.number_format = FMT_USD
    c.alignment = ALIGN_RIGHT
for col in (1, 2, 3, 4, 5):
    pd_.cell(row=r, column=col).fill = FILL_TOTAL
r += 2

pd_.cell(row=r, column=1, value="Timeline (months)").font = FONT_BODY
c = pd_.cell(row=r, column=3, value="=PreDev_TimelineMonths")
c.font = FONT_CROSSREF
c.number_format = FMT_NUM1
c.alignment = ALIGN_RIGHT

pd_.column_dimensions["A"].width = 46
pd_.column_dimensions["B"].width = 14
pd_.column_dimensions["C"].width = 14
pd_.column_dimensions["D"].width = 14
pd_.column_dimensions["E"].width = 16

# ============================================================
# Reorder tabs: README first
# ============================================================
order = ["README", "Inputs", "Path A - Micro", "Path B - Assemblage",
         "Capital Stack", "Waterfall v1", "Pre-Dev Budget"]
wb._sheets = [wb[name] for name in order]

# Save
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"Saved {OUT_PATH}")
print(f"Named ranges defined: {len(NAMES) + 2}")
